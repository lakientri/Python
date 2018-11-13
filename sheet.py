# Need to put the .py file and excel file within the same directory of openpyxl
# in this case is: C:\Users\TriLa\Anaconda3\Lib\site-packages
#https://www.youtube.com/watch?v=tJxcKyFMTGo
import  openpyxl 

wb=openpyxl.load_workbook("book1.xlsx")

#printing the sheet names
wb_sheet=wb.active
wb_sheet.title='Sheet1'
wb_sheet = wb.get_sheet_names()
print(wb_sheet)
#Change the sheet name - wb_sheet.title='Sheet'

#readVal=wb_sheet['A1'].value
#print(readVal)

# Read rows and column in excel
"""rows=4
column=3
for r in range(1,rows+1):
    for c in range(1, column+1):
        # format '%-8s'
        print(wb_sheet.cell(row=r, column=c).value, end=' ')
       # if((wb_sheet.cell(row=r, column=c).value) == "None")
    print('')"""

#wb.save("book1.xlsx")


